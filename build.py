# -*- coding: utf-8 -*-
r"""
Gera o arquivo dados.js (criptografado) a partir da planilha
"Fluxo de Caixa Fabrica.xlsx".

Uso:
    python build.py "C:\caminho\Fluxo de Caixa Fabrica.xlsx" MINHA_SENHA

O conteudo e cifrado com AES-256-GCM. A chave vem da senha via
PBKDF2-HMAC-SHA256 (310.000 iteracoes). Sem a senha o dados.js e
apenas ruido - mesmo com o repositorio publico.
"""
import sys, os, re, csv, json, base64, datetime, hashlib, unicodedata

import openpyxl
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

ITER = 310_000

# ------------------------------------------------------------------ config
# Arquivos com dado real (grupo.json, notas.json, dividas.csv, recebimentos.csv)
# moram no repositorio PRIVADO, nao aqui. Guardar copia dos dois lados criava
# divergencia silenciosa: editar um e o outro ficar velho sem ninguem notar.
#
# Ordem de busca:
#   1. variavel de ambiente ARTFLEX_CONFIG
#   2. repositorio privado clonado ao lado (config/)
#   3. esta pasta - fallback, para quem preferir copiar os arquivos
_AQUI = os.path.dirname(os.path.abspath(__file__))
_VIZINHOS = [
    os.path.join(os.path.dirname(_AQUI), "FINANCEIRO - ARTFLEX - CONTEXTO", "config"),
    os.path.join(os.path.dirname(_AQUI), "artflex-financeiro", "config"),
]


def cfg(nome):
    """Caminho do arquivo de configuracao, onde quer que ele esteja."""
    for base in [os.environ.get("ARTFLEX_CONFIG")] + _VIZINHOS + [_AQUI]:
        if base and os.path.exists(os.path.join(base, nome)):
            return os.path.join(base, nome)
    return os.path.join(_AQUI, nome)          # nao existe: o chamador trata


# ----------------------------------------------------------------- categorias
# ------------------------------------------------------------- exclusoes
# Lancamentos que a planilha traz mas que nao devem aparecer no painel,
# retirados a pedido. A planilha nao e alterada.
#
# O valor faz parte da chave de proposito: se a linha mudar de valor na
# planilha, ela volta a aparecer em vez de sumir sem ninguem perceber.
# O total retirado e mostrado no rodape e no diagnostico do painel - um corte
# de caixa invisivel seria pior do que nao cortar.
# Vazia: o painel mostra a planilha inteira. Com a lista vazia, o rodape e o
# cartao de exclusoes do diagnostico simplesmente nao aparecem.
EXCLUIR = []

# ------------------------------------------------------------------ grupo
# Entidades e contas vem de grupo.json, que fica SO na maquina de quem opera:
# regime tributario e relacao matriz/filial nao podem ficar em repositorio
# publico. O grupo-modelo.json versionado mostra o formato.
_G = {}
_gp = cfg("grupo.json")
if os.path.exists(_gp):
    with open(_gp, encoding="utf-8") as _f:
        _G = json.load(_f)
EMPRESAS = _G.get("empresas", {})
CONTAS = _G.get("contas", {})
CONTAS_ENCERRADAS = set(_G.get("contas_encerradas", []))
# Cadastros do ERP que nao devem entrar no de-para. Hoje sao dois fornecedores
# antigos que ficaram com o CNPJ da propria Artflex: entrariam como se a
# empresa fosse fornecedora de si mesma. O motivo de cada um esta no grupo.json.
CADASTROS_IGNORADOS = {(str(x.get("origem", "")), str(x.get("codigo", "")))
                       for x in _G.get("cadastros_ignorados", [])}


def _chave(s):
    """Normaliza para comparar: sem acento, sem pontuacao, maiusculo."""
    s = unicodedata.normalize("NFKD", (s or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^A-Z0-9]+", " ", s).strip()


def _chave_dura(s):
    """Chave que sobrevive ao acento perdido na exportacao do banco.

    O Santander exporta trocando a letra acentuada por espaco: "SEBASTIAO"
    chega como "SEBASTI O", "GONCALVES" como "GON ALVES", "MOVEIS" como
    "M VEIS". Tirar o acento nao resolve, porque a letra nao virou outra -
    ela sumiu.

    Entao esta chave APAGA a letra acentuada dos dois lados e joga fora todo
    separador. "SEBASTIAO" (com til) e "SEBASTI O" viram os dois SEBASTIO.

    E lossy de proposito, entao so serve de segunda tentativa, e so quando a
    chave leva a um unico CNPJ. Chave ambigua e descartada.
    """
    d = unicodedata.normalize("NFD", (s or "").strip().upper())
    fora, i = [], 0
    while i < len(d):
        if i + 1 < len(d) and unicodedata.combining(d[i + 1]):
            i += 1                                  # pula a letra de base
            while i < len(d) and unicodedata.combining(d[i]):
                i += 1                              # e as marcas dela
            continue
        fora.append(d[i])
        i += 1
    return re.sub(r"[^A-Z0-9]+", "", "".join(fora))


# Aceita o nome cru da planilha e tambem o de exibicao.
_APELIDOS = {}
for _k, _v in EMPRESAS.items():
    _APELIDOS[_chave(_k)] = _v
    _APELIDOS[_chave(_v["nome"])] = _v


def entidade(bruto):
    """Nome da planilha ou do CSV -> (nome de exibicao, bloco do grupo)."""
    e = _APELIDOS.get(_chave(bruto))
    if e:
        return e["nome"], e["grupo"]
    return ((bruto or "").strip() or "(sem empresa)"), "Não classificado"


# Pagamentos em que uma empresa paga conta de outra ("alguns pagamentos do
# varejo feito na industria"). Distorce o resultado por empresa. Aqui so
# marcamos CANDIDATOS pelo texto - a confirmacao e do Andriel, uma a uma.
INTERCO_PISTAS = ("SOFA OUTLET", "DECORART", "BESSA", "LOJA")

TRIBUTOS = {"PIS", "COFINS", "IPI", "IRPJ", "CSLL", "DIFAL", "ICMS DE VENDA",
            "FEEF", "SIMPLES", "INSS", "FGTS", "SERVICO SOCIAL", "ICMS"}
# "COMPENSADO PF" e fornecedor de materia-prima (compensado), pago pela PF.
FORNECEDORES = {"COMPENSADO PF", "TETRAOIL"}
UTILIDADES = {"CAGEPA", "ENERGISA", "BRISANET"}
# Seja AP e consultoria/treinamento; Softcom e Tek-System sao os ERPs.
SERVICOS = {"SOFTCOM", "TEK-SYSTEM", "TEKSYSTEM", "INFORPOINT", "SEGPOL",
            "BWA", "SEJA AP"}
MARKETING = {"FEIRA CARUARU", "TV SOFA OUTLET"}
# Maquinario comprado parcelado: e ativo, e a parcela e divida.
INVESTIMENTO = {"MAQUINA DE ESPUMA", "PLOTTER"}


def categoria(desc: str) -> str:
    d = (desc or "").upper().strip()
    if d.startswith("FORNECEDOR") or d in FORNECEDORES:
        return "Fornecedores e insumos"
    if d in TRIBUTOS:
        return "Tributos"
    if "ALUGUEL" in d or "GALP" in d:
        return "Aluguéis"
    if d in UTILIDADES:
        return "Utilidades"
    if d in SERVICOS:
        return "Serviços e sistemas"
    if d in MARKETING:
        return "Marketing"
    if d in INVESTIMENTO:
        return "Investimento / ativo"
    return "Outros"


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None



# ------------------------------------------------------------- recebimentos
# Porta de entrada do dinheiro que entra. Enquanto nao existir recebimentos.csv,
# o fluxo de 13 semanas mostra so o lado da saida - e diz isso na cara.
#
# Formato minimo, uma linha por titulo:
#     data;empresa;valor
# Colunas aceitas alem dessas: descricao, sacado, banco, documento.
# Datas em dd/mm/aaaa ou aaaa-mm-dd. Valores em 1.234,56 ou 1234.56.
COLUNAS = {
    "data": ("data", "vencimento", "data_vencimento", "dt_vencimento", "credito"),
    "empresa": ("empresa", "cnpj", "unidade", "filial"),
    "valor": ("valor", "vlr", "valor_titulo", "montante"),
    "descricao": ("descricao", "descrição", "historico", "histórico", "titulo"),
    "sacado": ("sacado", "cliente", "pagador", "contraparte"),
    "banco": ("banco", "carteira", "instituicao", "instituição"),
    "documento": ("documento", "doc", "nosso_numero", "titulo_numero", "nf"),
}


def _acha(cabecalho, alvos):
    norm = {re.sub(r"[^a-z_]", "", c.strip().lower().replace(" ", "_")): c
            for c in cabecalho}
    for a in alvos:
        if a in norm:
            return norm[a]
    return None


def _data(v):
    v = (v or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(v[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _valor(v):
    v = (v or "").strip().replace("R$", "").replace(" ", "")
    if not v:
        return None
    # "1.234,56" (pt-BR) vs "1234.56" (en): a virgula manda quando existe
    if "," in v:
        v = v.replace(".", "").replace(",", ".")
    try:
        return round(float(v), 2)
    except ValueError:
        return None


# --------------------------------------------------------------- recebiveis
# Carteira de cobranca exportada dos bancos, em recebiveis/*.xls|xlsx.
#
# A separacao que importa esta no nome do arquivo:
#   ...descontad...  -> titulo JA ANTECIPADO. O dinheiro entrou. Se o sacado
#                       nao pagar, a empresa RECOMPRA. Nao e entrada futura,
#                       e risco de recompra
#   qualquer outro   -> carteira simples. Titulo ainda a receber
#
# O banco sai da primeira palavra do nome. A empresa dona vem do mapa
# "carteiras" do grupo.json.
CARTEIRAS = _G.get("carteiras", {})


def _cnpj_raiz(doc):
    """Os 8 primeiros digitos. Matriz e filial do mesmo grupo compartilham."""
    d = re.sub(r"[^0-9]", "", doc or "")
    return d[:8] if len(d) >= 8 else d


def _linhas_xls(caminho):
    """Le .xls (xlrd) ou .xlsx (openpyxl) e devolve matriz de strings."""
    if caminho.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        return [["" if c is None else str(c).strip() for c in r]
                for r in ws.iter_rows(values_only=True)]
    import xlrd
    sh = xlrd.open_workbook(caminho).sheet_by_index(0)
    return [[str(sh.cell_value(i, j)).strip() for j in range(sh.ncols)]
            for i in range(sh.nrows)]


# Muitas fontes gravam o documento colado no nome do sacado:
#   Vector     -> "12.345.678/0001-99-NOME DO SACADO"   (CNPJ inteiro)
#   Credvale   -> "12.345.678 NOME DO SACADO"           (so a raiz)
_RE_CNPJ = re.compile(r"^\s*(\d{2}\.?\d{3}\.?\d{3})(?:/?\d{4}-?\d{2})?")


def _extrai_doc(texto):
    """Separa (documento, nome) quando o documento vem colado no nome."""
    m = _RE_CNPJ.match(texto or "")
    if not m:
        return "", (texto or "").strip()
    resto = (texto or "")[m.end():].lstrip(" -–")
    return m.group(0).strip(), resto.strip() or (texto or "").strip()


def _layout(linhas, nome):
    """Descobre banco, modalidade e a linha de cabecalho da tabela.

    O cabecalho e a primeira linha que traz, ao mesmo tempo, algo de
    vencimento e algo de sacado/pagador. Isso vale para os quatro layouts e
    nao quebra quando o banco mexe no numero de linhas do topo.
    """
    topo = _chave(" ".join(" ".join(r) for r in linhas[:10]))

    i_cab = 0
    for i, r in enumerate(linhas[:15]):
        k = _chave(" ".join(r))
        if "VENCIMENTO" in k and ("SACADO" in k or "PAGADOR" in k):
            i_cab = i
            break

    # FIDC: tudo que esta la ja foi cedido, entao e sempre descontado.
    if "CREDVALE" in topo:
        return "CREDVALE", "FIDC - CEDIDO", i_cab
    if "VECTOR" in topo:
        return "VECTOR", "FIDC - CEDIDO", i_cab

    # Santander grava a modalidade dentro do arquivo - melhor que o nome.
    if "TIPO COBRANCA" in topo or "COD BENEFICIARIO" in topo:
        modal = ""
        for i, r in enumerate(linhas[:8]):
            for j, c in enumerate(r):
                if "TIPO COBRAN" in _chave(c) and i + 1 < len(linhas):
                    if j < len(linhas[i + 1]):
                        modal = linhas[i + 1][j]
        return "SANTANDER", modal, i_cab

    banco = nome.split()[0].upper() if nome.split() else "?"
    return banco, ("DESCONTADO" if "DESCONTAD" in _chave(nome) else "SIMPLES"), i_cab


# Cadastro de pessoas do ERP Tek-System. E a saida definitiva para o sacado
# que chega sem CNPJ: em vez de adivinhar pelo nome parecido, pergunta ao
# cadastro onde o titulo nasceu.
#
# Como exportar (vale para Cliente e para Fornecedor):
#   Cadastros > Pessoas > {Cliente|Fornecedor} > botao Relatorio
#   > engrenagem > Modo de Saida > Analisar em Cubo > Previsao
#   > aba Dados Brutos > botao direito > Exportar Dados > Para Arquivo > CSV
#
# Sai em latin-1, separado por virgula.
_CAD_ARQUIVOS = [("clientes", "clientes-teksystem.csv"),
                 ("fornecedores", "fornecedores-teksystem.csv")]
_CAD_DOC = ["CNPJCEI_PESSOA_JUR", "CNPJ_PESSOA_END", "CPF_PESSOA_FIS"]
_CAD_NOME = ["RAZAOSOCIAL_PESSOA", "NOMEFANTASIA_PESSOA"]
AMBIGUO = object()          # sentinela: nome que leva a mais de uma raiz
# Prefixo menor que isso casa com gente demais para servir de prova.
_PREFIXO_MIN = 12


def ler_cadastro():
    """Indexa razao social e nome fantasia -> documento.

    Devolve (exato, duro, avisos). Os bancos gravam ora a razao social ora o
    nome fantasia, entao os dois entram no indice.

    A ambiguidade e julgada pela RAIZ do CNPJ, nao pelo documento inteiro.
    Matriz e filial sao cadastros diferentes com o mesmo nome - "GINALDO DA
    NOBREGA GONCALVES" aparece tres vezes, uma por filial. Para o de-para isso
    nao e ambiguidade nenhuma: os tres sao o mesmo risco de credito. So quando
    as raizes divergem e que a chave e jogada fora, e ai vai fora dos dois
    indices: se o cadastro nao sabe dizer quem e, o build nao pode fingir que
    sabe.

    A terceira saida e a lista crua (chave dura, raiz), usada para fechar o
    nome que o banco truncou.
    """
    exato, duro, lista, avisos = {}, {}, [], []
    lidos = ignorados = 0
    for origem, arquivo in _CAD_ARQUIVOS:
        caminho = cfg(arquivo)
        if not os.path.exists(caminho):
            continue
        linhas = []
        for enc in ("utf-8-sig", "latin-1"):
            try:
                with open(caminho, encoding=enc, newline="") as f:
                    linhas = list(csv.DictReader(f, delimiter=","))
                if linhas and len(linhas[0]) > 5:
                    break
            except Exception:
                continue
        if not linhas:
            avisos.append("cadastro: nao consegui ler %s" % arquivo)
            continue

        for r in linhas:
            cod = (r.get("CODIGO_PESSOA") or "").strip()
            if cod in ("", "0"):
                continue                          # linha "INDEFINIDA" do topo
            if (origem, cod) in CADASTROS_IGNORADOS:
                ignorados += 1
                continue
            doc = ""
            for c in _CAD_DOC:
                d = re.sub(r"[^0-9]", "", r.get(c) or "")
                if len(d) in (11, 14):
                    doc = d
                    break
            if not doc:
                continue                          # importador estrangeiro
            lidos += 1
            raiz = _cnpj_raiz(doc)
            for c in _CAD_NOME:
                nome = (r.get(c) or "").strip()
                if not nome:
                    continue
                kd = _chave_dura(nome)
                if kd:
                    lista.append((kd, raiz))
                for indice, chave in ((exato, _chave(nome)), (duro, kd)):
                    if not chave:
                        continue
                    antigo = indice.get(chave)
                    if antigo is None:
                        indice[chave] = (raiz, doc)
                    elif antigo == AMBIGUO:
                        pass
                    elif antigo[0] != raiz:
                        indice[chave] = AMBIGUO
                    elif antigo[1] != doc:
                        indice[chave] = (raiz, "")   # mesma raiz, filial outra

    for indice in (exato, duro):
        for k in [k for k, v in indice.items() if v is AMBIGUO]:
            del indice[k]

    if lidos:
        avisos.append("cadastro: %d pessoas do ERP com documento, %d nomes "
                      "indexados" % (lidos, len(exato)))
    if ignorados:
        avisos.append("cadastro: %d cadastros ignorados por configuracao"
                      % ignorados)
    return exato, duro, lista, avisos


def ler_recebiveis(pasta):
    base = os.path.normpath(os.path.join(
        os.path.dirname(cfg("grupo.json")), "..", "recebiveis"))
    if not os.path.isdir(base):
        base = os.path.join(pasta, "recebiveis")
    if not os.path.isdir(base):
        return [], []

    titulos, avisos = [], []
    for nome in sorted(os.listdir(base)):
        if not nome.lower().endswith((".xls", ".xlsx")) or nome.startswith("~$"):
            continue
        try:
            linhas = _linhas_xls(os.path.join(base, nome))
        except Exception as e:
            avisos.append("recebiveis: nao consegui abrir %s (%s)" % (nome, e))
            continue

        banco, modal, i_cab = _layout(linhas, nome)
        carteira = "simples" if _chave(modal).startswith("SIMPLES") or \
            ("DESC" not in _chave(modal) and "CEDIDO" not in _chave(modal)) else "descontada"
        empresa_bruta = CARTEIRAS.get(banco)
        if not empresa_bruta:
            avisos.append("recebiveis: banco %r sem empresa no mapa 'carteiras' "
                          "do grupo.json - %s ignorado" % (banco, nome))
            continue
        nome_emp, grupo_emp = entidade(empresa_bruta)

        cab = [_chave(c) for c in linhas[i_cab]]

        def col(*alvos):
            for a in alvos:
                for j, c in enumerate(cab):
                    if a in c:
                        return j
            return None

        c_nome = col("NOME DO PAGADOR", "PAGADOR", "SACADO", "CLIENTE")
        c_doc = col("CPF CNPJ", "CNPJ", "DOCUMENTO DO")
        c_venc = col("VENCIMENTO")
        # "VALOR ABERTO" vem antes de "VALOR DE FACE": e o saldo, nao o original
        c_val = col("VALOR ABERTO", "VLR ATUAL", "VALOR DO TITULO", "VALOR DE FACE", "VALOR")
        c_nn = col("NOSSO N")
        c_sn = col("SEU N", "N DOCUMENTO", "DOCUMENTO")
        c_sit = col("SITUA", "STATUS")
        c_emi = col("EMISS")
        c_atr = col("ATRASO", "ATRS")
        if c_venc is None or c_val is None:
            avisos.append("recebiveis: %s sem coluna de vencimento ou valor "
                          "(cabecalho na linha %d)" % (nome, i_cab + 1))
            continue

        lidos = 0
        for r in linhas[i_cab + 1:]:
            if len(r) <= max(c_venc, c_val):
                continue
            venc = _data(r[c_venc])
            valor = _valor(r[c_val])
            if not venc or not valor:
                continue                       # total, rodape, legenda ou vazio
            pega = lambda j: (r[j].strip() if j is not None and j < len(r) else "")
            doc = pega(c_doc)
            sacado = pega(c_nome)
            if not doc:                        # documento colado no nome?
                doc, sacado = _extrai_doc(sacado)
            titulos.append({
                "d": venc, "v": valor, "e": nome_emp, "g": grupo_emp,
                "carteira": carteira, "banco": banco, "modalidade": modal,
                "sacado": sacado, "cnpj": doc, "raiz": _cnpj_raiz(doc),
                "nn": pega(c_nn), "sn": pega(c_sn), "sit": pega(c_sit),
                "emissao": _data(pega(c_emi)), "atraso": pega(c_atr),
                "origem": nome,
            })
            lidos += 1
        if not lidos:
            avisos.append("recebiveis: %s nao rendeu nenhum titulo" % nome)

    # Sacado que chegou sem CNPJ: tres tentativas, todas por igualdade exata
    # depois de normalizar. Nenhuma delas e por semelhanca.
    #
    # Aproximacao por semelhanca foi testada em 25/08/2026 e reprovada: dos 4
    # pares que ela apontou, todos os que o Andriel conferiu eram clientes
    # DIFERENTES - um par com o mesmo sobrenome em empresas sem relacao, outro
    # com a mesma razao social separada so por iniciais no comeco. Os dois
    # foram apontados com 100% de confianca.
    #
    # Casar dois clientes errados nao da erro nenhum: so deixa o numero errado
    # para sempre. Melhor ficar sem identificar do que identificar errado.
    for t in titulos:
        t["fonte_raiz"] = "arquivo" if t["raiz"] else ""

    # 1. Outro titulo, de qualquer carteira, com o mesmo nome e com CNPJ.
    por_nome = {}
    for t in titulos:
        if t["raiz"]:
            por_nome.setdefault(_chave(t["sacado"]), t["raiz"])
    for t in titulos:
        if not t["raiz"]:
            r = por_nome.get(_chave(t["sacado"]))
            if r:
                t["raiz"], t["fonte_raiz"] = r, "outro titulo"

    # 2 e 3. O cadastro do ERP - primeiro pelo nome exato, depois pela chave
    # que tolera o acento que o Santander comeu.
    cad_exato, cad_duro, cad_lista, cad_avisos = ler_cadastro()
    avisos.extend(cad_avisos)
    for t in titulos:
        if t["raiz"]:
            continue
        achado = cad_exato.get(_chave(t["sacado"]))
        fonte = "cadastro do ERP"
        if not achado:
            achado = cad_duro.get(_chave_dura(t["sacado"]))
            fonte = "cadastro do ERP (acento perdido)"
        if achado:
            raiz, doc = achado
            t["raiz"], t["fonte_raiz"] = raiz, fonte
            if doc:
                t["cnpj"] = doc

    # 4. Nome truncado pelo banco. O Santander e a Credvale cortam em ~40
    # caracteres, entao "NOGUEIRA COMERCIO VAREJISTA DE MOVEIS LT" nunca vai
    # bater com "...LTDA" por igualdade.
    #
    # Isto continua NAO sendo semelhanca: exige que o nome do banco seja o
    # comeco literal do nome do cadastro, e que todos os cadastros que comecam
    # assim tenham a MESMA raiz. Se duas empresas diferentes compartilham o
    # prefixo, ninguem e escolhido.
    for t in titulos:
        if t["raiz"]:
            continue
        k = _chave_dura(t["sacado"])
        if len(k) < _PREFIXO_MIN:
            continue
        raizes = {r for kd, r in cad_lista if kd.startswith(k)}
        if len(raizes) == 1:
            t["raiz"] = raizes.pop()
            t["fonte_raiz"] = "cadastro do ERP (nome truncado)"

    conta = {}
    for t in titulos:
        if t["raiz"] and t["fonte_raiz"] != "arquivo":
            conta[t["fonte_raiz"]] = conta.get(t["fonte_raiz"], 0) + 1
    for fonte in sorted(conta, key=lambda k: -conta[k]):
        avisos.append("recebiveis: %d titulos identificados por %s"
                      % (conta[fonte], fonte))
    sem = [t for t in titulos if not t["raiz"]]
    if sem:
        nomes = sorted({t["sacado"] for t in sem})
        avisos.append("recebiveis: %d titulos de %d sacados seguem sem CNPJ" %
                      (len(sem), len(nomes)))

    titulos.sort(key=lambda t: (t["d"], t["sacado"]))
    return titulos, avisos


def ler_recebimentos(pasta):
    caminho = cfg("recebimentos.csv")
    if not os.path.exists(caminho):
        return [], []

    with open(caminho, encoding="utf-8-sig", newline="") as f:
        amostra = f.read(4096)
        f.seek(0)
        try:
            sep = csv.Sniffer().sniff(amostra, delimiters=";,\t").delimiter
        except csv.Error:
            sep = ";"
        linhas = list(csv.DictReader(f, delimiter=sep))

    if not linhas:
        return [], ["recebimentos.csv esta vazio"]

    cab = list(linhas[0].keys())
    col = {k: _acha(cab, v) for k, v in COLUNAS.items()}
    faltando = [k for k in ("data", "empresa", "valor") if not col[k]]
    if faltando:
        return [], ["recebimentos.csv sem as colunas: " + ", ".join(faltando) +
                    " (achei: " + ", ".join(cab) + ")"]

    recebimentos, avisos = [], []
    for n, ln in enumerate(linhas, start=2):
        data, valor = _data(ln.get(col["data"])), _valor(ln.get(col["valor"]))
        if not data or valor is None:
            avisos.append(f"linha {n} ignorada: data ou valor invalido")
            continue
        nome, grupo = entidade(ln.get(col["empresa"]))
        pega = lambda k: (ln.get(col[k]) or "").strip() if col[k] else ""
        recebimentos.append({
            "d": data, "e": nome, "g": grupo, "v": valor,
            "n": pega("descricao") or "Recebimento",
            "s": pega("sacado"), "b": pega("banco"), "doc": pega("documento"),
        })
    return recebimentos, avisos


# ------------------------------------------------------------------- dividas
# Emprestimos e linhas rotativas, em dividas.csv (fica so na maquina do Andriel;
# o modelo versionado tem dado ficticio). As parcelas NAO estao na planilha, e
# por isso o fluxo de caixa vinha subestimando a saida todo mes.
#
# tipo:
#   sac         - amortiza capital fixo por mes; juros incidem sobre o saldo
#   rotativo    - so juros sobre o saldo, sem amortizacao (conta garantida)
#   fixa        - parcela constante informada em capital_mes
#   antecipacao - duplicata cedida a FIDC com coobrigacao. Nao gera parcela: o
#                 sacado e quem paga. Mas o saldo e RISCO DE RECOMPRA - se o
#                 sacado nao pagar, a Artflex recompra o titulo. Por isso entra
#                 no quadro de divida, separado das operacoes que amortizam.
def ler_dividas(pasta):
    caminho = cfg("dividas.csv")
    if not os.path.exists(caminho):
        return [], []
    with open(caminho, encoding="utf-8-sig", newline="") as f:
        linhas = list(csv.DictReader(f, delimiter=";"))

    dividas, avisos = [], []
    for n, ln in enumerate(linhas, start=2):
        try:
            saldo = _valor(ln.get("saldo"))
            taxa = _valor(ln.get("taxa_am"))
            cap = _valor(ln.get("capital_mes")) or 0.0
            dia = int((ln.get("dia") or "1").strip() or 1)
        except (TypeError, ValueError):
            avisos.append(f"dividas.csv linha {n}: numero invalido")
            continue
        if saldo is None or taxa is None:
            avisos.append(f"dividas.csv linha {n}: saldo ou taxa ausente")
            continue
        nome, grupo = entidade(ln.get("empresa"))
        dividas.append({
            "op": (ln.get("operacao") or "").strip(),
            "e": nome, "g": grupo,
            "banco": (ln.get("banco") or "").strip(),
            "produto": (ln.get("produto") or "").strip(),
            "tipo": (ln.get("tipo") or "sac").strip().lower(),
            "saldo": saldo, "taxa": taxa, "cap": cap, "dia": dia,
            "fim": _data(ln.get("fim")),
            "obs": (ln.get("obs") or "").strip(),
            # so para antecipacao: face cedido, quanto entrou de fato, quanto
            # voltou como recompra, e a taxa anunciada na capa do bordero
            "face": _valor(ln.get("face")),
            "liquido": _valor(ln.get("liquido")),
            "recompra": _valor(ln.get("recompra")),
            "taxa_capa": _valor(ln.get("taxa_capa")),
        })
    return dividas, avisos


def projetar_parcelas(dividas, ate_meses=6):
    """Gera as parcelas futuras de cada operacao, no mesmo formato dos
    pagamentos da planilha, para entrarem no fluxo de caixa."""
    hoje = datetime.date.today()
    parcelas = []
    for d in dividas:
        if d["tipo"] == "antecipacao":
            continue          # quem paga e o sacado; nao ha parcela nossa
        saldo = d["saldo"]
        fim = datetime.date.fromisoformat(d["fim"]) if d["fim"] else None
        ref = hoje.replace(day=1)
        for _ in range(ate_meses + 1):
            ano, mes = ref.year, ref.month
            ultimo = [31, 29 if ano % 4 == 0 and (ano % 100 or ano % 400 == 0) else 28,
                      31, 30, 31, 30, 31, 31, 30, 31, 30, 31][mes - 1]
            venc = datetime.date(ano, mes, min(d["dia"], ultimo))
            ref = datetime.date(ano + (mes == 12), (mes % 12) + 1, 1)
            if venc < hoje or (fim and venc > fim) or saldo <= 0.01:
                continue
            juros = round(saldo * d["taxa"] / 100, 2)
            amort = 0.0 if d["tipo"] == "rotativo" else min(d["cap"], saldo)
            valor = round(juros + amort, 2) if d["tipo"] != "fixa" else d["cap"]
            if valor <= 0:
                continue
            parcelas.append({
                "d": venc.isoformat(), "e": d["e"], "g": d["g"], "x": 0,
                "n": f"{d['banco']} · {d['produto']}",
                "v": valor, "c": "Dívida bancária", "l": None,
                "juros": juros, "amort": round(amort, 2), "op": d["op"],
            })
            saldo = round(saldo - amort, 2)
    parcelas.sort(key=lambda p: (p["d"], p["e"]))
    return parcelas


# ------------------------------------------------------------------- extracao
def extrair(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

    # --- aba SEMANAL: agenda de pagamentos -------------------------------
    ws = wb["SEMANAL"]
    linhas = list(ws.iter_rows(min_row=1, max_row=400, values_only=True))

    pagamentos, subtotais, excluidos = [], [], []
    for i, r in enumerate(linhas, start=1):
        if i < 4:
            continue
        desc, dt, val, emp = r[5], r[6], r[7], r[8]
        if not any(x is not None and str(x).strip() for x in (desc, dt, val, emp)):
            continue
        # linha de subtotal semanal: numero solto na coluna da descricao
        if isinstance(desc, (int, float)) and dt is None and val is None:
            subtotais.append(round(float(desc), 2))
            continue
        data = dt.date().isoformat() if isinstance(dt, (datetime.datetime, datetime.date)) else None
        if not data:
            continue
        d = txt(desc) or "(sem descricao)"
        nome, grupo = entidade(txt(emp))
        # candidato a intercompany: a industria pagando algo que nomeia uma loja
        interco = grupo == "Indústria" and any(p in d.upper() for p in INTERCO_PISTAS)
        reg = {
            "d": data,                       # data
            "e": nome,                       # empresa (nome de exibicao)
            "g": grupo,                      # bloco: Industria / Varejo / Socio
            "x": 1 if interco else 0,        # candidato a operacao entre empresas
            "n": d,                          # descricao / natureza
            "v": num(val),                   # valor (None = ainda sem valor)
            "c": categoria(d),               # categoria inferida
            "l": i,                          # linha de origem na planilha
        }
        if any((reg["d"], reg["e"], reg["n"]) == (dt_, e_, n_)
               and abs((reg["v"] or 0) - v) < 0.005
               for dt_, e_, n_, v in EXCLUIR):
            excluidos.append(reg)
            continue
        pagamentos.append(reg)

    # --- saldos bancarios (colunas K/L = atual, N/O = anterior) ----------
    saldos = []
    for i, r in enumerate(linhas, start=1):
        if not (3 <= i <= 13):
            continue
        conta, atual, ant = txt(r[10]), num(r[11]), num(r[14])
        if conta and conta.upper() != "TOTAL":
            dono = CONTAS.get(conta.upper())
            saldos.append({
                "conta": conta, "atual": atual, "anterior": ant,
                "e": dono or "(sem dono)",
                "g": (entidade(dono)[1] if dono else "Não classificado"),
                "off": 1 if conta.upper() in CONTAS_ENCERRADAS else 0,
            })

    # --- avisos fixos do cabecalho (linha 1) -----------------------------
    fixos = [txt(v) for v in linhas[0][2:] if txt(v)]

    # A aba TOTAL nao e lida de proposito: seus numeros estao desatualizados
    # (confirmado com o Andriel em 24/08/2026 - o bloco de fornecedores marca
    # desatualizados). A SEMANAL e a
    # unica fonte do painel. Para voltar a usar a TOTAL, ler as colunas A/B
    # (boletos), F/G/H (pagamentos do mes) e J/K (saldos).

    # Valores que a propria planilha exibe (ultimo calculo salvo pelo Excel).
    # Guardados para poder citar o numero que o usuario ve, sem recalcular.
    resumo = {
        "semanal": {"receitas": num(linhas[15][11]), "despesas": num(linhas[16][11]),
                    "resultado": num(linhas[17][11]), "saldo_total": num(linhas[12][11])},
    }

    pasta = os.path.dirname(os.path.abspath(__file__))
    recebimentos, avisos_rec = ler_recebimentos(pasta)
    titulos, avisos_tit = ler_recebiveis(pasta)
    # Carteira simples e entrada futura. Descontada NAO e: o dinheiro ja
    # entrou, e o que sobra e o risco de recompra no vencimento.
    for t in titulos:
        if t['carteira'] == 'simples':
            recebimentos.append({'d': t['d'], 'e': t['e'], 'g': t['g'],
                                 'v': t['v'], 'n': 'Boleto ' + t['banco'],
                                 's': t['sacado'], 'b': t['banco'], 'doc': t['nn']})

    # notas.json: analise em prosa que NAO pode ficar no codigo, porque o
    # repositorio e publico. Entra cifrada no dados.js, como o resto.
    notas = []
    _n = cfg("notas.json")
    if os.path.exists(_n):
        with open(_n, encoding="utf-8") as f:
            notas = json.load(f)
    dividas, avisos_div = ler_dividas(pasta)
    parcelas = projetar_parcelas(dividas)

    # Tabela de dimensao das entidades - so as que aparecem nos dados.
    usadas = {p["e"] for p in pagamentos} | {r["e"] for r in recebimentos}
    dim_empresas = [{"nome": v["nome"], "grupo": v["grupo"],
                     "regime": v["regime"], "cnpj": v["cnpj"]}
                    for v in EMPRESAS.values() if v["nome"] in usadas]

    return {
        "pagamentos": pagamentos,
        "recebimentos": recebimentos,
        "titulos": titulos,
        "avisos_titulos": avisos_tit,
        # Parcelas de emprestimo ficam FORA de "pagamentos" de proposito: elas
        # nao vem da planilha e nao podem entrar na conferencia dos subtotais.
        # O fluxo de caixa soma as duas listas.
        "dividas": dividas,
        "parcelas": parcelas,
        "avisos_dividas": avisos_div,
        "empresas": dim_empresas,
        "notas": notas,
        "avisos_recebimentos": avisos_rec,
        "excluidos": excluidos,
        "subtotais_semanais": subtotais,
        "saldos": saldos,
        "fixos": fixos,
        "resumo": resumo,
        "hoje": datetime.date.today().isoformat(),
        "gerado_em": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
        "origem": os.path.basename(caminho),
    }


# ---------------------------------------------------------------- criptografia
def cifrar(dados: dict, senha: str) -> str:
    claro = json.dumps(dados, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    salt = os.urandom(16)
    iv = os.urandom(12)
    chave = hashlib.pbkdf2_hmac("sha256", senha.encode("utf-8"), salt, ITER, 32)
    ct = AESGCM(chave).encrypt(iv, claro, None)
    b64 = lambda b: base64.b64encode(b).decode("ascii")
    pacote = {"v": 1, "it": ITER, "salt": b64(salt), "iv": b64(iv), "ct": b64(ct)}
    return "window.__ARTFLEX__=" + json.dumps(pacote) + ";\n"


def carimbar(pasta):
    """Poe ?v=<hash> nas tags <script> do index.html.

    O GitHub Pages entrega os arquivos com cache de alguns minutos. Sem esse
    carimbo, um navegador que ja visitou a pagina continua rodando o app.js
    antigo - e foi assim que, depois de uma troca de senha, a senha nova
    aparecia como incorreta. O hash muda so quando app.js ou dados.js mudam,
    entao o index.html nao fica sujo a toa.
    """
    h = hashlib.sha256()
    for nome in ("app.js", "dados.js"):
        with open(os.path.join(pasta, nome), "rb") as f:
            h.update(f.read())
    versao = h.hexdigest()[:8]

    idx = os.path.join(pasta, "index.html")
    with open(idx, encoding="utf-8") as f:
        html = f.read()
    novo = re.sub(r'src="(dados|app)\.js(?:\?v=[0-9a-f]+)?"',
                  lambda m: 'src="%s.js?v=%s"' % (m.group(1), versao), html)
    if novo != html:
        with open(idx, "w", encoding="utf-8") as f:
            f.write(novo)
    return versao


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    caminho, senha = sys.argv[1], sys.argv[2]
    dados = extrair(caminho)

    com_valor = [p for p in dados["pagamentos"] if p["v"]]
    total = sum(p["v"] for p in com_valor)
    pasta = os.path.dirname(os.path.abspath(__file__))
    destino = os.path.join(pasta, "dados.js")
    with open(destino, "w", encoding="utf-8") as f:
        f.write(cifrar(dados, senha))
    versao = carimbar(pasta)

    def brl(v):
        return ("R$ %s" % f"{v:,.2f}").replace(",", "X").replace(".", ",").replace("X", ".")

    excl = sum(p["v"] or 0 for p in dados["excluidos"])
    subtotais = sum(dados["subtotais_semanais"])

    rec = dados["recebimentos"]
    print(f"OK  {len(dados['pagamentos'])} pagamentos "
          f"({len(com_valor)} com valor, "
          f"{len(dados['pagamentos']) - len(com_valor)} sem valor)")
    print(f"    total no painel        {brl(total)}")
    if rec:
        print(f"    recebimentos           {brl(sum(r['v'] for r in rec))}  "
              f"({len(rec)} titulos)")
    else:
        print("    recebimentos           nenhum - crie recebimentos.csv "
              "(modelo em recebimentos-modelo.csv)")
    for a in (dados["avisos_recebimentos"] + dados["avisos_dividas"]
              + dados["avisos_titulos"]):
        print(f"    !! {a}")
    if dados["dividas"]:
        banc = [d for d in dados["dividas"] if d["tipo"] != "antecipacao"]
        ante = [d for d in dados["dividas"] if d["tipo"] == "antecipacao"]
        sd = sum(d["saldo"] for d in banc)
        jm = sum(d["saldo"] * d["taxa"] / 100 for d in banc)
        print(f"    divida bancaria        {brl(sd)}  ({len(banc)} operacoes)")
        print(f"    juros/mes              {brl(jm)}  = {brl(jm * 12)}/ano")
        if ante:
            sa = sum(d["saldo"] for d in ante)
            print(f"    titulos cedidos        {brl(sa)}  ({len(ante)} borderos) "
                  f"- risco de recompra, nao e parcela")
        p6 = dados["parcelas"]
        print(f"    parcelas projetadas    {len(p6)} nos proximos 6 meses, "
              f"{brl(sum(p['v'] for p in p6))}")
        print("    custo do capital, do mais barato ao mais caro:")
        for d in sorted(dados["dividas"], key=lambda x: x["taxa"]):
            aa = ((1 + d["taxa"] / 100) ** 12 - 1) * 100
            print(f"      {d['taxa']:>5.3f}% a.m.  {aa:>6.2f}% a.a.   "
                  f"{d['banco']:<10} {d['produto'][:34]:<34} {brl(d['saldo'])}")
    interco = [p for p in dados["pagamentos"] if p.get("x")]
    if interco:
        print(f"    entre empresas         {len(interco)} candidatos, "
              f"{brl(sum(p['v'] or 0 for p in interco))} (conferir)")
    if dados["excluidos"]:
        print(f"    retirados a pedido     {brl(excl)}  ({len(dados['excluidos'])} lancamentos)")
        for p in dados["excluidos"]:
            print(f"      - {p['d']}  {p['e']:<12} {p['n']:<18} {brl(p['v'] or 0)}")
    # A conferencia so fecha somando de volta o que foi retirado: e o que prova
    # que a extracao nao perdeu nenhuma outra linha pelo caminho.
    print(f"    conferencia            {brl(total + excl)} vs "
          f"{brl(subtotais)} de subtotais da planilha"
          f"  {'OK' if abs(total + excl - subtotais) < 0.01 else '<<< NAO BATE'}")
    print(f"    -> {destino}")
    print(f"    index.html carimbado com ?v={versao}"
          f"  (commite index.html junto com dados.js)")


if __name__ == "__main__":
    main()
